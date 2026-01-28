"""Utility functions - Import/Export Excel"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.http import HttpResponse
from io import BytesIO
from .models import Bieu1KQ2025, Bieu2KH2026, Bieu3KH20262030, Bieu4CTThanhPho, Ward

def import_all_bieu_donvi(ward, excel_file):
    """Import tất cả 3 biểu từ 1 file Excel có 3 sheet
    Tìm sheet dựa vào tên, không phụ thuộc vào vị trí
    """
    try:
        results = []
        wb = openpyxl.load_workbook(excel_file)
        
        # Tìm sheet cho từng biểu - kiểm tra chính xác
        bieu1_sheet = None
        bieu2_sheet = None
        bieu3_sheet = None
        
        for sheet_name in wb.sheetnames:
            name_lower = sheet_name.lower()
            
            # Biểu 1: có 'kq' hoặc có '01'
            if 'kq' in name_lower or '01' in name_lower:
                bieu1_sheet = wb[sheet_name]
            # Biểu 2: có '02' nhưng không chứa '03'
            if '02' in name_lower and '03' not in name_lower:
                bieu2_sheet = wb[sheet_name]
            # Biểu 3: có '03' hoặc có '2026-2030'
            if '03' in name_lower or '2026-2030' in name_lower:
                bieu3_sheet = wb[sheet_name]
        
        # Biểu 1
        if bieu1_sheet:
            count1 = _import_bieu1_from_sheet(ward, bieu1_sheet)
            results.append(f"Biểu 1: {count1} trường")
        else:
            results.append("Biểu 1: 0 trường")
        
        # Biểu 2
        if bieu2_sheet:
            count2 = _import_bieu2_from_sheet(ward, bieu2_sheet)
            results.append(f"Biểu 2: {count2} trường")
        else:
            results.append("Biểu 2: 0 trường")
        
        # Biểu 3
        if bieu3_sheet:
            count3 = _import_bieu3_from_sheet(ward, bieu3_sheet)
            results.append(f"Biểu 3: {count3} trường")
        else:
            results.append("Biểu 3: 0 trường")
        
        return True, f"Đã import thành công: {', '.join(results)}"
    except Exception as e:
        return False, f"Lỗi: {str(e)}"

def _import_bieu1_from_sheet(ward, ws):
    """Import Biểu 1 từ worksheet"""
    imported_count = 0
    
    # Tìm vị trí các marker I và II
    row_ii_marker = None
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        if cell_a and str(cell_a).strip().upper() == 'II' and row_ii_marker is None:
            row_ii_marker = row_idx
    
    def get_loai_cong_nhan(row_idx):
        if row_ii_marker and row_idx >= row_ii_marker:
            return 'CN_LAI'
        else:
            return 'CN_MOI'
    
    for row_idx in range(4, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        
        if cell_a and str(cell_a).strip().upper() in ['I', 'II']:
            continue
        
        if not cell_b or str(cell_b).strip() == '':
            continue
        
        ten_truong = str(cell_b).strip()
        cap_hoc = ws.cell(row=row_idx, column=3).value
        loai_hinh = ws.cell(row=row_idx, column=4).value
        nam_dat_chuan = ws.cell(row=row_idx, column=5).value
        muc_do_cqg = ws.cell(row=row_idx, column=6).value
        so_quyet_dinh = ws.cell(row=row_idx, column=7).value
        ghi_chu = ws.cell(row=row_idx, column=8).value
        loai_cong_nhan = get_loai_cong_nhan(row_idx)
        
        Bieu1KQ2025.objects.update_or_create(
            ward=ward,
            ten_truong=ten_truong,
            defaults={
                'cap_hoc': str(cap_hoc).strip() if cap_hoc else '',
                'loai_hinh': str(loai_hinh).strip() if loai_hinh else '',
                'nam_dat_chuan_gan_nhat': str(nam_dat_chuan).strip() if nam_dat_chuan and str(nam_dat_chuan).strip() != '' else '',
                'muc_do_cqg': str(muc_do_cqg).strip() if muc_do_cqg else '',
                'so_quyet_dinh_cqg': str(so_quyet_dinh).strip() if so_quyet_dinh else '',
                'loai_cong_nhan': loai_cong_nhan,
                'ghi_chu': str(ghi_chu).strip() if ghi_chu else '',
            }
        )
        imported_count += 1
    
    return imported_count

def _import_bieu2_from_sheet(ward, ws):
    """Import Biểu 2 từ worksheet"""
    imported_count = 0
    
    # Tìm vị trí các marker I và II
    row_ii_marker = None
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        if cell_a and str(cell_a).strip().upper() == 'II' and row_ii_marker is None:
            row_ii_marker = row_idx
    
    def get_loai_cong_nhan(row_idx):
        if row_ii_marker and row_idx >= row_ii_marker:
            return 'CN_LAI'
        else:
            return 'CN_MOI'
    
    for row_idx in range(4, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        
        if cell_a and str(cell_a).strip().upper() in ['I', 'II']:
            continue
        
        if not cell_b or str(cell_b).strip() == '':
            continue
        
        ten_truong = str(cell_b).strip()
        cap_hoc = ws.cell(row=row_idx, column=3).value
        loai_hinh = ws.cell(row=row_idx, column=4).value
        nam_dat_cqg = ws.cell(row=row_idx, column=5).value
        da_kiem_tra = ws.cell(row=row_idx, column=6).value
        du_kien_thang = ws.cell(row=row_idx, column=7).value
        ghi_chu = ws.cell(row=row_idx, column=8).value
        loai_cong_nhan = get_loai_cong_nhan(row_idx)
        
        Bieu2KH2026.objects.update_or_create(
            ward=ward,
            ten_truong=ten_truong,
            defaults={
                'cap_hoc': str(cap_hoc).strip() if cap_hoc else '',
                'loai_hinh': str(loai_hinh).strip() if loai_hinh else '',
                'nam_dat_cqg_gan_nhat': str(nam_dat_cqg).strip() if nam_dat_cqg else '',
                'phuong_xa_da_kiem_tra': str(da_kiem_tra).strip() if da_kiem_tra else '',
                'du_kien_thang': str(du_kien_thang).strip() if du_kien_thang else '',
                'loai_cong_nhan': loai_cong_nhan,
                'ghi_chu': str(ghi_chu).strip() if ghi_chu else '',
            }
        )
        imported_count += 1
    
    return imported_count

def _import_bieu3_from_sheet(ward, ws):
    """Import Biểu 3 từ worksheet"""
    imported_count = 0
    
    # Tìm vị trí các marker I và II
    row_ii_marker = None
    for row_idx in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        if cell_a and str(cell_a).strip().upper() == 'II' and row_ii_marker is None:
            row_ii_marker = row_idx
    
    def get_loai_cong_nhan(row_idx):
        if row_ii_marker and row_idx >= row_ii_marker:
            return 'CN_LAI'
        else:
            return 'CN_MOI'
    
    for row_idx in range(4, ws.max_row + 1):
        cell_a = ws.cell(row=row_idx, column=1).value
        cell_b = ws.cell(row=row_idx, column=2).value
        
        if cell_a and str(cell_a).strip().upper() in ['I', 'II']:
            continue
        
        if not cell_b or str(cell_b).strip() == '':
            continue
        
        ten_truong = str(cell_b).strip()
        cap_hoc = ws.cell(row=row_idx, column=3).value
        loai_hinh = ws.cell(row=row_idx, column=4).value
        nam_dat_cqg = ws.cell(row=row_idx, column=5).value
        
        cn_moi_2026 = ws.cell(row=row_idx, column=6).value
        cn_moi_2027 = ws.cell(row=row_idx, column=7).value
        cn_moi_2028 = ws.cell(row=row_idx, column=8).value
        cn_moi_2029 = ws.cell(row=row_idx, column=9).value
        cn_moi_2030 = ws.cell(row=row_idx, column=10).value
        
        cn_lai_2026 = ws.cell(row=row_idx, column=11).value
        cn_lai_2027 = ws.cell(row=row_idx, column=12).value
        cn_lai_2028 = ws.cell(row=row_idx, column=13).value
        cn_lai_2029 = ws.cell(row=row_idx, column=14).value
        cn_lai_2030 = ws.cell(row=row_idx, column=15).value
        
        ghi_chu = ws.cell(row=row_idx, column=16).value
        
        Bieu3KH20262030.objects.update_or_create(
            ward=ward,
            ten_truong=ten_truong,
            defaults={
                'cap_hoc': str(cap_hoc).strip() if cap_hoc else '',
                'loai_hinh': str(loai_hinh).strip() if loai_hinh else '',
                'nam_dat_cqg_gan_nhat': str(nam_dat_cqg).strip() if nam_dat_cqg else '',
                'cn_moi_2026': str(cn_moi_2026).strip() if cn_moi_2026 else '',
                'cn_moi_2027': str(cn_moi_2027).strip() if cn_moi_2027 else '',
                'cn_moi_2028': str(cn_moi_2028).strip() if cn_moi_2028 else '',
                'cn_moi_2029': str(cn_moi_2029).strip() if cn_moi_2029 else '',
                'cn_moi_2030': str(cn_moi_2030).strip() if cn_moi_2030 else '',
                'cn_lai_2026': str(cn_lai_2026).strip() if cn_lai_2026 else '',
                'cn_lai_2027': str(cn_lai_2027).strip() if cn_lai_2027 else '',
                'cn_lai_2028': str(cn_lai_2028).strip() if cn_lai_2028 else '',
                'cn_lai_2029': str(cn_lai_2029).strip() if cn_lai_2029 else '',
                'cn_lai_2030': str(cn_lai_2030).strip() if cn_lai_2030 else '',
                'ghi_chu': str(ghi_chu).strip() if ghi_chu else '',
            }
        )
        imported_count += 1
    
    return imported_count

def import_bieu1_donvi(ward, excel_file):
    """Import Biểu 1 từ file donvi-sample.xlsx
    Tìm marker I (CN mới) và II (CN lại) để phân loại dữ liệu
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        imported_count = 0
        
        # Tìm vị trí các marker I và II
        row_i_marker = None
        row_ii_marker = None
        
        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value
            if cell_a and str(cell_a).strip().upper() == 'I' and row_i_marker is None:
                row_i_marker = row_idx
            elif cell_a and str(cell_a).strip().upper() == 'II' and row_ii_marker is None:
                row_ii_marker = row_idx
        
        def get_loai_cong_nhan(row_idx):
            if row_ii_marker and row_idx >= row_ii_marker:
                return 'CN_LAI'
            else:
                return 'CN_MOI'
        
        for row_idx in range(4, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_b = ws.cell(row=row_idx, column=2).value
            
            # Bỏ qua hàng marker I/II
            if cell_a and str(cell_a).strip().upper() in ['I', 'II']:
                continue
            
            if not cell_b or str(cell_b).strip() == '':
                continue
            
            ten_truong = str(cell_b).strip()
            cap_hoc = ws.cell(row=row_idx, column=3).value
            loai_hinh = ws.cell(row=row_idx, column=4).value
            nam_dat_chuan = ws.cell(row=row_idx, column=5).value
            muc_do_cqg = ws.cell(row=row_idx, column=6).value
            so_quyet_dinh = ws.cell(row=row_idx, column=7).value
            ghi_chu = ws.cell(row=row_idx, column=8).value
            loai_cong_nhan = get_loai_cong_nhan(row_idx)
            
            Bieu1KQ2025.objects.update_or_create(
                ward=ward,
                ten_truong=ten_truong,
                defaults={
                    'cap_hoc': str(cap_hoc).strip() if cap_hoc else '',
                    'loai_hinh': str(loai_hinh).strip() if loai_hinh else '',
                    'nam_dat_chuan_gan_nhat': str(nam_dat_chuan).strip() if nam_dat_chuan and str(nam_dat_chuan).strip() != '' else '',
                    'muc_do_cqg': str(muc_do_cqg).strip() if muc_do_cqg else '',
                    'so_quyet_dinh_cqg': str(so_quyet_dinh).strip() if so_quyet_dinh else '',
                    'loai_cong_nhan': loai_cong_nhan,
                    'ghi_chu': str(ghi_chu).strip() if ghi_chu else '',
                }
            )
            imported_count += 1
        
        return True, f"Đã import {imported_count} trường"
    except Exception as e:
        return False, f"Lỗi: {str(e)}"

def import_bieu2_donvi(ward, excel_file):
    """Import Biểu 2 từ file donvi-sample.xlsx
    Cấu trúc: Tìm marker I (CN mới) và II (CN lại) để phân loại dữ liệu
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        imported_count = 0
        
        # Tìm vị trí các marker I và II
        row_i_marker = None  # Row chứa "I" - bắt đầu CN mới
        row_ii_marker = None  # Row chứa "II" - bắt đầu CN lại
        
        for row_idx in range(1, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value
            if cell_a and str(cell_a).strip().upper() == 'I' and row_i_marker is None:
                row_i_marker = row_idx
            elif cell_a and str(cell_a).strip().upper() == 'II' and row_ii_marker is None:
                row_ii_marker = row_idx
        
        # Xác định loại công nhận dựa vào vị trí row
        def get_loai_cong_nhan(row_idx):
            if row_ii_marker and row_idx >= row_ii_marker:
                return 'CN_LAI'
            else:
                return 'CN_MOI'
        
        for row_idx in range(4, ws.max_row + 1):  # Bắt đầu từ row 4
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_b = ws.cell(row=row_idx, column=2).value
            
            # Bỏ qua hàng marker I/II
            if cell_a and str(cell_a).strip().upper() in ['I', 'II']:
                continue
            
            if not cell_b or str(cell_b).strip() == '':
                continue
            
            ten_truong = str(cell_b).strip()
            cap_hoc = ws.cell(row=row_idx, column=3).value
            loai_hinh = ws.cell(row=row_idx, column=4).value
            nam_dat_cqg = ws.cell(row=row_idx, column=5).value
            da_kiem_tra = ws.cell(row=row_idx, column=6).value
            du_kien_thang = ws.cell(row=row_idx, column=7).value
            ghi_chu = ws.cell(row=row_idx, column=8).value
            loai_cong_nhan = get_loai_cong_nhan(row_idx)
            
            Bieu2KH2026.objects.update_or_create(
                ward=ward,
                ten_truong=ten_truong,
                defaults={
                    'cap_hoc': str(cap_hoc).strip() if cap_hoc else '',
                    'loai_hinh': str(loai_hinh).strip() if loai_hinh else '',
                    'nam_dat_cqg_gan_nhat': str(nam_dat_cqg).strip() if nam_dat_cqg else '',
                    'phuong_xa_da_kiem_tra': str(da_kiem_tra).strip() if da_kiem_tra else '',
                    'du_kien_thang': str(du_kien_thang).strip() if du_kien_thang else '',
                    'loai_cong_nhan': loai_cong_nhan,
                    'ghi_chu': str(ghi_chu).strip() if ghi_chu else '',
                }
            )
            imported_count += 1
        
        return True, f"Đã import {imported_count} trường"
    except Exception as e:
        return False, f"Lỗi: {str(e)}"

def import_bieu3_donvi(ward, excel_file):
    """Import Biểu 3 từ file donvi-sample.xlsx
    Cấu trúc: Cột 1-5: STT, Tên trường, Cấp học, Loại hình, Năm đạt CQG
              Cột 6-10: CN mới 2026-2030
              Cột 11-15: CN lại 2026-2030
              Cột 16: Ghi chú
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        imported_count = 0
        
        for row_idx in range(4, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_b = ws.cell(row=row_idx, column=2).value
            
            # Bỏ qua hàng phân loại I/II
            if cell_a and str(cell_a).strip().upper() in ['I', 'II']:
                continue
            
            # Bỏ qua hàng rỗng
            if not cell_b or str(cell_b).strip() == '':
                continue
            
            ten_truong = str(cell_b).strip()
            cap_hoc = ws.cell(row=row_idx, column=3).value
            loai_hinh = ws.cell(row=row_idx, column=4).value
            nam_dat_cqg = ws.cell(row=row_idx, column=5).value
            
            # Đọc CN mới (cột 6-10: năm 2026-2030)
            cn_moi_2026 = ws.cell(row=row_idx, column=6).value
            cn_moi_2027 = ws.cell(row=row_idx, column=7).value
            cn_moi_2028 = ws.cell(row=row_idx, column=8).value
            cn_moi_2029 = ws.cell(row=row_idx, column=9).value
            cn_moi_2030 = ws.cell(row=row_idx, column=10).value
            
            # Đọc CN lại (cột 11-15: năm 2026-2030)
            cn_lai_2026 = ws.cell(row=row_idx, column=11).value
            cn_lai_2027 = ws.cell(row=row_idx, column=12).value
            cn_lai_2028 = ws.cell(row=row_idx, column=13).value
            cn_lai_2029 = ws.cell(row=row_idx, column=14).value
            cn_lai_2030 = ws.cell(row=row_idx, column=15).value
            
            ghi_chu = ws.cell(row=row_idx, column=16).value
            
            Bieu3KH20262030.objects.update_or_create(
                ward=ward,
                ten_truong=ten_truong,
                defaults={
                    'cap_hoc': str(cap_hoc).strip() if cap_hoc else '',
                    'loai_hinh': str(loai_hinh).strip() if loai_hinh else '',
                    'nam_dat_cqg_gan_nhat': str(nam_dat_cqg).strip() if nam_dat_cqg else '',
                    'cn_moi_2026': str(cn_moi_2026).strip() if cn_moi_2026 else '',
                    'cn_moi_2027': str(cn_moi_2027).strip() if cn_moi_2027 else '',
                    'cn_moi_2028': str(cn_moi_2028).strip() if cn_moi_2028 else '',
                    'cn_moi_2029': str(cn_moi_2029).strip() if cn_moi_2029 else '',
                    'cn_moi_2030': str(cn_moi_2030).strip() if cn_moi_2030 else '',
                    'cn_lai_2026': str(cn_lai_2026).strip() if cn_lai_2026 else '',
                    'cn_lai_2027': str(cn_lai_2027).strip() if cn_lai_2027 else '',
                    'cn_lai_2028': str(cn_lai_2028).strip() if cn_lai_2028 else '',
                    'cn_lai_2029': str(cn_lai_2029).strip() if cn_lai_2029 else '',
                    'cn_lai_2030': str(cn_lai_2030).strip() if cn_lai_2030 else '',
                    'ghi_chu': str(ghi_chu).strip() if ghi_chu else '',
                }
            )
            imported_count += 1
        
        return True, f"Đã import {imported_count} trường"
    except Exception as e:
        return False, f"Lỗi: {str(e)}"

def export_bieu1_tonghop():
    """Export Biểu 1 tổng hợp - Tất cả trường từ tất cả phường/xã"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biểu 1"
    
    # Header
    ws.merge_cells('A1:I1')
    ws['A1'] = 'TỔNG HỢP KẾT QUẢ CÔNG NHẬN MỚI và CÔNG NHẬN LẠI TRƯỜNG CHUẨN QUỐC GIA NĂM 2025'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:I2')
    ws['A2'] = '(Đính kèm công văn số        /UBND-     , ngày    /01/2026 của UBND ...)'
    
    # Column headers
    headers = ['STT', 'Quận, Huyện', 'Phường, xã', 'Tên trường đạt chuẩn quốc gia', 
               'Cấp học', 'Loại hình', 'Năm đạt chuẩn', 'Mức độ CQG', 'Ghi chú']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Data - Công nhận mới
    row = 6
    ws.cell(row=row, column=3, value='I. CÔNG NHẬN MỚI').font = Font(bold=True)
    row += 1
    
    # Lấy tất cả phường/xã
    all_wards = Ward.objects.all().order_by('stt')
    
    # Lặp qua từng phường/xã
    for ward in all_wards:
        records_cn_moi = Bieu1KQ2025.objects.filter(
            ward=ward, 
            loai_cong_nhan='CN_MOI'
        ).order_by('ten_truong')
        
        if records_cn_moi.exists():
            # Có dữ liệu - xuất từng trường
            for record in records_cn_moi:
                ws.cell(row=row, column=1, value=record.ward.stt)
                ws.cell(row=row, column=3, value=record.ward.don_vi)
                ws.cell(row=row, column=4, value=record.ten_truong)
                ws.cell(row=row, column=5, value=record.cap_hoc)
                ws.cell(row=row, column=6, value=record.loai_hinh)
                ws.cell(row=row, column=7, value=record.nam_dat_chuan_gan_nhat)
                ws.cell(row=row, column=8, value=record.muc_do_cqg)
                ws.cell(row=row, column=9, value=record.ghi_chu)
                row += 1
        else:
            # Không có dữ liệu - tạo dòng trống
            ws.cell(row=row, column=1, value=ward.stt)
            ws.cell(row=row, column=3, value=ward.don_vi)
            ws.cell(row=row, column=9, value='Không có')
            row += 1
    
    # Data - Công nhận lại
    row += 1
    ws.cell(row=row, column=3, value='II. CÔNG NHẬN LẠI').font = Font(bold=True)
    row += 1
    
    # Lặp qua từng phường/xã
    for ward in all_wards:
        records_cn_lai = Bieu1KQ2025.objects.filter(
            ward=ward,
            loai_cong_nhan='CN_LAI'
        ).order_by('ten_truong')
        
        if records_cn_lai.exists():
            # Có dữ liệu - xuất từng trường
            for record in records_cn_lai:
                ws.cell(row=row, column=1, value=record.ward.stt)
                ws.cell(row=row, column=3, value=record.ward.don_vi)
                ws.cell(row=row, column=4, value=record.ten_truong)
                ws.cell(row=row, column=5, value=record.cap_hoc)
                ws.cell(row=row, column=6, value=record.loai_hinh)
                ws.cell(row=row, column=7, value=record.nam_dat_chuan_gan_nhat)
                ws.cell(row=row, column=8, value=record.muc_do_cqg)
                ws.cell(row=row, column=9, value=record.ghi_chu)
                row += 1
        else:
            # Không có dữ liệu - tạo dòng trống
            ws.cell(row=row, column=1, value=ward.stt)
            ws.cell(row=row, column=3, value=ward.don_vi)
            ws.cell(row=row, column=9, value='Không có')
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    
    # Save to response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="bieu1-tonghop.xlsx"'
    return response

def export_bieu2_tonghop():
    """Export Biểu 2 tổng hợp"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biểu 2"
    
    ws.merge_cells('A1:H1')
    ws['A1'] = 'TỔNG HỢP KẾ HOẠCH CÔNG NHẬN MỚI VÀ CÔNG NHẬN LẠI NĂM 2026'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    headers = ['STT', 'Phường/xã', 'Tên trường', 'Cấp học', 'Loại hình', 
               'Năm đạt CQG', 'Đã kiểm tra', 'Ghi chú']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
    
    # Data - Công nhận mới
    row = 5
    ws.cell(row=row, column=2, value='I. CÔNG NHẬN MỚI').font = Font(bold=True)
    row += 1
    
    # Lấy tất cả phường/xã
    all_wards = Ward.objects.all().order_by('stt')
    
    # Lặp qua từng phường/xã cho CN mới
    for ward in all_wards:
        records_cn_moi = Bieu2KH2026.objects.filter(
            ward=ward,
            loai_cong_nhan='CN_MOI'
        ).order_by('ten_truong')
        
        if records_cn_moi.exists():
            # Có dữ liệu - xuất từng trường
            for record in records_cn_moi:
                ws.cell(row=row, column=1, value=record.ward.stt)
                ws.cell(row=row, column=2, value=record.ward.don_vi)
                ws.cell(row=row, column=3, value=record.ten_truong)
                ws.cell(row=row, column=4, value=record.cap_hoc)
                ws.cell(row=row, column=5, value=record.loai_hinh)
                ws.cell(row=row, column=6, value=record.nam_dat_cqg_gan_nhat)
                ws.cell(row=row, column=7, value=record.phuong_xa_da_kiem_tra)
                ws.cell(row=row, column=8, value=record.ghi_chu)
                row += 1
        else:
            # Không có dữ liệu - tạo dòng trống
            ws.cell(row=row, column=1, value=ward.stt)
            ws.cell(row=row, column=2, value=ward.don_vi)
            ws.cell(row=row, column=8, value='Không có')
            row += 1
    
    # Data - Công nhận lại
    row += 1
    ws.cell(row=row, column=2, value='II. CÔNG NHẬN LẠI').font = Font(bold=True)
    row += 1
    
    # Lặp qua từng phường/xã cho CN lại
    for ward in all_wards:
        records_cn_lai = Bieu2KH2026.objects.filter(
            ward=ward,
            loai_cong_nhan='CN_LAI'
        ).order_by('ten_truong')
        
        if records_cn_lai.exists():
            # Có dữ liệu - xuất từng trường
            for record in records_cn_lai:
                ws.cell(row=row, column=1, value=record.ward.stt)
                ws.cell(row=row, column=2, value=record.ward.don_vi)
                ws.cell(row=row, column=3, value=record.ten_truong)
                ws.cell(row=row, column=4, value=record.cap_hoc)
                ws.cell(row=row, column=5, value=record.loai_hinh)
                ws.cell(row=row, column=6, value=record.nam_dat_cqg_gan_nhat)
                ws.cell(row=row, column=7, value=record.phuong_xa_da_kiem_tra)
                ws.cell(row=row, column=8, value=record.ghi_chu)
                row += 1
        else:
            # Không có dữ liệu - tạo dòng trống
            ws.cell(row=row, column=1, value=ward.stt)
            ws.cell(row=row, column=2, value=ward.don_vi)
            ws.cell(row=row, column=8, value='Không có')
            row += 1
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bieu2-tonghop.xlsx"'
    return response

def export_bieu3_tonghop():
    """Export Biểu 3 tổng hợp theo đúng mẫu"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biểu 3"
    
    # Header chính
    ws.merge_cells('A1:Q1')
    ws['A1'] = 'KẾ HOẠCH CÔNG NHẬN MỚI VÀ CÔNG NHẬN LẠI TRƯỜNG CHUẨN QUỐC GIA GIAI ĐOẠN 2026-2030'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers dòng 2
    ws.cell(row=2, column=1, value='STT')
    ws.cell(row=2, column=2, value='Phường, xã')
    ws.cell(row=2, column=3, value='Tên trường trong kế hoạch')
    ws.cell(row=2, column=4, value='Cấp học')
    ws.cell(row=2, column=5, value='Loại hình')
    ws.cell(row=2, column=6, value='Năm đạt CQG gần nhất')
    
    # Merge cells cho Công nhận mới
    ws.merge_cells('G2:K2')
    ws['G2'] = 'Công nhận mới'
    ws['G2'].font = Font(bold=True)
    ws['G2'].alignment = Alignment(horizontal='center')
    
    # Merge cells cho Công nhận lại
    ws.merge_cells('L2:P2')
    ws['L2'] = 'Công nhận lại'
    ws['L2'].font = Font(bold=True)
    ws['L2'].alignment = Alignment(horizontal='center')
    
    # Cột ghi chú
    ws.cell(row=2, column=17, value='Ghi chú')
    ws.cell(row=2, column=17).font = Font(bold=True)
    
    # Headers dòng 3 - Năm
    for col, year in enumerate(['2026', '2027', '2028', '2029', '2030'], 7):
        ws.cell(row=3, column=col, value=f'Năm {year}')
        ws.cell(row=3, column=col).font = Font(bold=True)
    for col, year in enumerate(['2026', '2027', '2028', '2029', '2030'], 12):
        ws.cell(row=3, column=col, value=f'Năm {year}')
        ws.cell(row=3, column=col).font = Font(bold=True)
    
    # Data
    row = 4
    all_wards = Ward.objects.all().order_by('stt')
    
    for ward in all_wards:
        records = Bieu3KH20262030.objects.filter(ward=ward).order_by('ten_truong')
        
        if records.exists():
            # Có dữ liệu - xuất từng trường
            for record in records:
                ws.cell(row=row, column=1, value=record.ward.stt)
                ws.cell(row=row, column=2, value=record.ward.don_vi)
                ws.cell(row=row, column=3, value=record.ten_truong)
                ws.cell(row=row, column=4, value=record.cap_hoc)
                ws.cell(row=row, column=5, value=record.loai_hinh)
                ws.cell(row=row, column=6, value=record.nam_dat_cqg_gan_nhat)
                # CN mới
                ws.cell(row=row, column=7, value=record.cn_moi_2026 or '')
                ws.cell(row=row, column=8, value=record.cn_moi_2027 or '')
                ws.cell(row=row, column=9, value=record.cn_moi_2028 or '')
                ws.cell(row=row, column=10, value=record.cn_moi_2029 or '')
                ws.cell(row=row, column=11, value=record.cn_moi_2030 or '')
                # CN lại
                ws.cell(row=row, column=12, value=record.cn_lai_2026 or '')
                ws.cell(row=row, column=13, value=record.cn_lai_2027 or '')
                ws.cell(row=row, column=14, value=record.cn_lai_2028 or '')
                ws.cell(row=row, column=15, value=record.cn_lai_2029 or '')
                ws.cell(row=row, column=16, value=record.cn_lai_2030 or '')
                ws.cell(row=row, column=17, value=record.ghi_chu or '')
                row += 1
        else:
            # Không có dữ liệu - tạo dòng trống
            ws.cell(row=row, column=1, value=ward.stt)
            ws.cell(row=row, column=2, value=ward.don_vi)
            ws.cell(row=row, column=17, value='Không có')
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    for col in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
        ws.column_dimensions[col].width = 10
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bieu3-tonghop.xlsx"'
    return response

def export_bieu4_tonghop():
    """Export Biểu 4 - Tự động tính từ Biểu 1"""
    # Tự động tạo/cập nhật Bieu4 từ dữ liệu Bieu1
    from django.db.models import Count, Q
    
    # Xóa dữ liệu cũ
    Bieu4CTThanhPho.objects.all().delete()
    
    # Tạo dữ liệu mới từ Bieu1 cho tất cả phường/xã
    for ward in Ward.objects.all().order_by('stt'):
        cn_moi = Bieu1KQ2025.objects.filter(ward=ward, loai_cong_nhan='CN_MOI').count()
        cn_lai = Bieu1KQ2025.objects.filter(ward=ward, loai_cong_nhan='CN_LAI').count()
        
        # Tạo record cho tất cả phường/xã, kể cả khi không có dữ liệu
        ghi_chu = 'Không có' if (cn_moi == 0 and cn_lai == 0) else ''
        
        Bieu4CTThanhPho.objects.create(
            ward=ward,
            stt=ward.stt,
            don_vi=ward.don_vi,
            cong_nhan_moi=cn_moi,
            cong_nhan_lai=cn_lai,
            ghi_chu=ghi_chu,
            loai=ward.loai
        )
    
    # Export Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biểu 4"
    
    ws['A1'] = 'TT'
    ws['B1'] = 'Đơn vị'
    ws['C1'] = 'Công nhận mới'
    ws['D1'] = 'Công nhận lại'
    ws['E1'] = 'Ghi chú'
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    row = 2
    records = Bieu4CTThanhPho.objects.all().order_by('stt')
    for record in records:
        ws.cell(row=row, column=1, value=record.stt)
        ws.cell(row=row, column=2, value=record.don_vi)
        ws.cell(row=row, column=3, value=record.cong_nhan_moi)
        ws.cell(row=row, column=4, value=record.cong_nhan_lai)
        ws.cell(row=row, column=5, value=record.ghi_chu or '')
        row += 1
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bieu4-tonghop.xlsx"'
    return response
