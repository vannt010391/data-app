import pandas as pd
import openpyxl

print("=== BIEU 1 DONVI SAMPLE ===")
wb = openpyxl.load_workbook('../bieu1-donvi-sample.xlsx')
ws = wb.active
print(f'Sheet name: {ws.title}')
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}')
print()

for i in range(1, min(30, ws.max_row+1)):
    row = [str(cell.value) if cell.value is not None else '' for cell in ws[i]]
    print(f'Row {i}: {row}')

print("\n\n=== BIEU 1 TONG HOP SAMPLE ===")
wb2 = openpyxl.load_workbook('../bieu1-tonghop-sample.xlsx')
ws2 = wb2.active
print(f'Sheet name: {ws2.title}')
print(f'Max row: {ws2.max_row}, Max col: {ws2.max_column}')
print()

for i in range(1, min(12, ws2.max_row+1)):
    row = [str(cell.value) if cell.value is not None else '' for cell in ws2[i]]
    print(f'Row {i}: {row}')

print("\n\n=== BIEU 2 TONG HOP SAMPLE ===")
wb3 = openpyxl.load_workbook('../bieu2-tonghop-sample.xlsx')
ws3 = wb3.active
print(f'Sheet name: {ws3.title}')
print(f'Max row: {ws3.max_row}, Max col: {ws3.max_column}')
for i in range(1, min(12, ws3.max_row+1)):
    row = [str(cell.value) if cell.value is not None else '' for cell in ws3[i]]
    print(f'Row {i}: {row}')

print("\n\n=== BIEU 3 TONG HOP SAMPLE ===")
wb4 = openpyxl.load_workbook('../bieu3-tonghop-sample.xlsx')
ws4 = wb4.active
print(f'Sheet name: {ws4.title}')
print(f'Max row: {ws4.max_row}, Max col: {ws4.max_column}')
for i in range(1, min(12, ws4.max_row+1)):
    row = [str(cell.value) if cell.value is not None else '' for cell in ws4[i]]
    print(f'Row {i}: {row}')

print("\n\n=== BIEU 4 TONG HOP UPDATE ===")
wb5 = openpyxl.load_workbook('../bieu4-tonghop-update.xlsx')
ws5 = wb5.active
print(f'Sheet name: {ws5.title}')
print(f'Max row: {ws5.max_row}, Max col: {ws5.max_column}')
for i in range(1, min(20, ws5.max_row+1)):
    row = [str(cell.value) if cell.value is not None else '' for cell in ws5[i]]
    print(f'Row {i}: {row}')
